from openpyxl import load_workbook

inv_file = load_workbook("inventory.xlsx")
product_list = inv_file['Sheet1']


def get_how_many_product_offer_by_each_supplier():
    product_per_supplier = {}
    for product_row in range(2, product_list.max_row + 1):
        supplier_name = product_list.cell(product_row, 4).value
        if supplier_name not in product_per_supplier:
            product_per_supplier[supplier_name] = 1
        else:
            product_per_supplier[supplier_name] += 1
    # have_most_product_company_name = max(product_per_supplier, key=product_per_supplier.get)
    # print(f"Have most product company is [{have_most_product_company_name}], and they have \
    # {product_per_supplier.get(have_most_product_company_name)} products~!")
    return product_per_supplier


def get_how_many_inventory_each_supplier():
    product_inventory_per_supplier = {}
    for product_row in range(2, product_list.max_row + 1):
        current_supplier_name = product_list.cell(product_row, 4).value
        current_product_inventory = int(product_list.cell(product_row, 2).value)
        if current_supplier_name in product_inventory_per_supplier:
            product_inventory_per_supplier[current_supplier_name] += current_product_inventory
        else:
            product_inventory_per_supplier[current_supplier_name] = current_product_inventory
    return product_inventory_per_supplier


def get_inventory_total_value_each_supplier():
    total_inventory_value_per_supplier = {}
    for product_row in range(2, product_list.max_row + 1):
        current_supplier_name = product_list.cell(product_row, 4).value
        current_product_inventory = int(product_list.cell(product_row, 2).value)
        current_product_value = float(product_list.cell(product_row, 3).value)

        if current_supplier_name in total_inventory_value_per_supplier:
            total_inventory_value_per_supplier[current_supplier_name] += current_product_inventory * current_product_value
        else:
            total_inventory_value_per_supplier[current_supplier_name] = current_product_inventory * current_product_value

        # logic add a column for all inventory price
        current_product_inventory_price = current_product_inventory * current_product_value
        product_list.cell(product_row, 5).value = current_product_inventory_price
    inv_file.save("inventory_with_total_value.xlsx")

    return total_inventory_value_per_supplier


if __name__ == '__main__':
    # num_of_product_supply_by_each_supplier = get_how_many_product_offer_by_each_supplier()
    # print(num_of_product_supply_by_each_supplier)
    #
    # inventory_per_supplier = get_how_many_inventory_each_supplier()
    # print(inventory_per_supplier)

    invertory_value_per_supplier = get_inventory_total_value_each_supplier()
    print(invertory_value_per_supplier)
